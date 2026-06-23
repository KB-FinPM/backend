from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

import app.db.base  # noqa: F401
from app.agents.core_agents.wbs_agent.agent import WbsAgent
from app.agents.output_agents.chat_agent.agent import ChatOutputAgent
from app.dependencies import (
    get_document_service,
    get_input_orchestrator,
    get_output_orchestrator,
)
from app.schemas.agent import AgentRequest
from app.schemas.artifact import DocumentMetadata, DocumentType
from app.schemas.chat import ChatActionCommand, ChatCommandType, ChatMessageRequest
from app.schemas.io_agent import (
    InputAgentResponse,
    NormalizedRequestType,
    OutputAgentResponse,
)
from app.services.artifact_export_service import ArtifactExportService


PROJECT_ID = "PRJ-TEST-001"
PROJECT_NAME = "KB Star Banking Process"
INTERNAL_FRAGMENTS = (
    "Input Agent",
    "Output Agent",
    "GenerationOrchestrator",
    "GENERATE_ARTIFACT",
    "structured_context",
    "traceback",
)


def assert_user_facing_message(message: str) -> None:
    assert message.strip()
    lowered = message.lower()
    for fragment in INTERNAL_FRAGMENTS:
        assert fragment.lower() not in lowered


def confirm_request(conversation_id: str, action_id: str) -> ChatMessageRequest:
    return ChatMessageRequest(
        project_id=PROJECT_ID,
        conversation_id=conversation_id,
        user_id="USER-001",
        message="confirm",
        action=ChatActionCommand(
            type=ChatCommandType.CONFIRM_PENDING_ACTION,
            action_id=action_id,
            payload={"action_id": action_id},
        ),
    )


@pytest.mark.anyio
async def test_user_flow_requests_requirement_spec_then_gets_upload_prompt(
    scenario_services,
) -> None:
    orchestrator = scenario_services["chat_orchestrator"]

    response = await orchestrator.handle_message(
        ChatMessageRequest(
            project_id=PROJECT_ID,
            user_id="USER-001",
            message="요구사항 정의서 만들어줘",
        )
    )

    assert response.state == "WAITING_REQUIRED_INFO"
    assert response.pending_action is None
    assert response.result["upload_request"]["required"] is True
    assert response.result["upload_request"]["documentType"] == (
        "CONSTRUCTION_REQUIREMENT_DEFINITION"
    )
    assert response.result["upload_request"]["acceptedTypes"]
    assert_user_facing_message(response.message)


@pytest.mark.anyio
async def test_user_flow_asks_requirement_spec_question_without_generation(
    scenario_services,
) -> None:
    orchestrator = scenario_services["chat_orchestrator"]
    generation_service = scenario_services["generation_service"]

    response = await orchestrator.handle_message(
        ChatMessageRequest(
            project_id=PROJECT_ID,
            user_id="USER-001",
            message="요구사항 정의서가 뭐야?",
        )
    )

    assert response.state == "IDLE"
    assert response.pending_action is None
    assert response.download_files == []
    assert "upload_request" not in response.result
    assert generation_service.requests == []
    assert_user_facing_message(response.message)


class UploadDocumentService:
    def __init__(self) -> None:
        self.received_document_id: str | None = None
        self.received_parsed_context: dict[str, Any] | None = None

    async def upload_to_storage(
        self,
        *,
        file_bytes: bytes,
        project_id: str,
        document_id: str,
        file_name: str,
        upload_prefix: str,
    ) -> str:
        return f"mock://upload/{project_id}/{document_id}/{file_name}"

    async def ingest_uploaded_document(
        self,
        *,
        document_id: str,
        project_id: str,
        document_type: DocumentType,
        file_name: str,
        storage_path: str,
        file_bytes: bytes,
        parsed_context: dict | None = None,
    ) -> DocumentMetadata:
        self.received_document_id = document_id
        self.received_parsed_context = parsed_context
        return DocumentMetadata(
            document_id=document_id,
            project_id=project_id,
            document_type=document_type,
            file_name=file_name,
            storage_path=storage_path,
        )


class UploadInputOrchestrator:
    async def normalize(self, request):
        return InputAgentResponse(
            agent_name="UploadInputOrchestrator",
            normalized_request_type=NormalizedRequestType.DOCUMENT_INGESTION,
            structured_context={
                "text": request.files[0].file_bytes.decode("utf-8"),
                "requirements": [
                    {
                        "requirement_id": "REQ-001",
                        "title": "Mobile login",
                    }
                ],
                "parser_name": "E2EParserStub",
            },
        )


class UploadOutputOrchestrator:
    async def format(self, request):
        return OutputAgentResponse(
            agent_name="UploadOutputOrchestrator",
            display_payload={
                "message": "document uploaded",
                "document": request.result_json["document"],
            },
        )


def test_user_flow_uploads_rfp_then_receives_document_id(
    client: TestClient,
) -> None:
    document_service = UploadDocumentService()
    client.app.dependency_overrides[get_document_service] = lambda: document_service
    client.app.dependency_overrides[get_input_orchestrator] = (
        lambda: UploadInputOrchestrator()
    )
    client.app.dependency_overrides[get_output_orchestrator] = (
        lambda: UploadOutputOrchestrator()
    )

    try:
        response = client.post(
            "/api/upload",
            data={
                "project_id": PROJECT_ID,
                "document_type": "CONSTRUCTION_REQUIREMENT_DEFINITION",
            },
            files={
                "file": (
                    "kb-rfp.txt",
                    (
                        "Project KB Star Banking Process\n"
                        "REQ-001. Users can sign in on mobile.\n"
                    ).encode("utf-8"),
                    "text/plain",
                )
            },
        )
    finally:
        client.app.dependency_overrides.clear()

    body = response.json()
    assert response.status_code == 200
    assert body["document"]["document_id"].startswith("DOC-")
    assert body["document"]["file_name"] == "kb-rfp.txt"
    assert body["document"]["document_type"] == "CONSTRUCTION_REQUIREMENT_DEFINITION"
    assert body["document"]["storage_path"].startswith("mock://upload/")
    assert document_service.received_document_id == body["document"]["document_id"]
    assert document_service.received_parsed_context["parser_name"] == "E2EParserStub"


@pytest.mark.anyio
async def test_user_flow_uploads_rfp_then_generates_requirement_spec(
    scenario_services,
) -> None:
    orchestrator = scenario_services["chat_orchestrator"]
    generation_service = scenario_services["generation_service"]

    first_response = await orchestrator.handle_message(
        ChatMessageRequest(
            project_id=PROJECT_ID,
            user_id="USER-001",
            message="방금 올린 문서 기준으로 요구사항 정의서 만들어줘",
            context={
                "project_name": PROJECT_NAME,
                "selected_document_ids": ["DOC-RFP-001"],
                "selected_documents": [
                    {
                        "document_id": "DOC-RFP-001",
                        "file_name": "kb-rfp.txt",
                        "display_label": "uploaded RFP",
                    }
                ],
                "source_document_type": "CONSTRUCTION_REQUIREMENT_DEFINITION",
            },
        )
    )

    assert first_response.state == "WAITING_CONFIRMATION"
    assert first_response.pending_action is not None
    assert_user_facing_message(first_response.message)

    completed_response = await orchestrator.handle_message(
        confirm_request(
            first_response.conversation_id,
            first_response.pending_action.action_id,
        )
    )

    assert completed_response.state == "COMPLETED"
    assert completed_response.download_files
    assert completed_response.download_files[0]["artifact_id"].startswith(
        "ART-REQUIREMENT_SPEC"
    )
    assert generation_service.requests[-1].target_artifact_type == "REQUIREMENT_SPEC"
    assert generation_service.requests[-1].source_document_ids == ["DOC-RFP-001"]
    assert_user_facing_message(completed_response.message)


@pytest.mark.anyio
async def test_user_flow_generates_wbs_from_requirement_spec(
    scenario_services,
) -> None:
    orchestrator = scenario_services["chat_orchestrator"]
    generation_service = scenario_services["generation_service"]

    first_response = await orchestrator.handle_message(
        ChatMessageRequest(
            project_id=PROJECT_ID,
            user_id="USER-001",
            message="요구사항 정의서를 기준으로 WBS 만들어줘",
            context={
                "project_name": PROJECT_NAME,
                "selected_document_ids": ["DOC-REQ-001"],
                "selected_documents": [
                    {
                        "document_id": "DOC-REQ-001",
                        "file_name": "kb-requirement-spec.xlsx",
                        "display_label": "requirement spec",
                    }
                ],
                "source_document_type": "REQUIREMENT_SPEC",
            },
        )
    )

    assert first_response.state == "WAITING_CONFIRMATION"
    assert first_response.pending_action is not None

    completed_response = await orchestrator.handle_message(
        confirm_request(
            first_response.conversation_id,
            first_response.pending_action.action_id,
        )
    )

    assert completed_response.state == "COMPLETED"
    assert completed_response.download_files
    assert completed_response.download_files[0]["artifact_id"].startswith("ART-WBS")
    assert generation_service.requests[-1].target_artifact_type == "WBS"
    assert generation_service.requests[-1].source_document_ids == ["DOC-REQ-001"]
    assert_user_facing_message(completed_response.message)


@pytest.mark.anyio
async def test_user_flow_generates_wbs_with_project_schedule_dates() -> None:
    response = await WbsAgent().generate(
        AgentRequest(
            project_id=PROJECT_ID,
            documents=[{"chunk_id": "CHUNK-001", "text": "mobile login account list"}],
            context={
                "project_name": PROJECT_NAME,
                "start_date": "2024.01.10",
                "project_period": "6개월",
                "requirement_artifact": {
                    "requirements": [
                        {
                            "requirement_id": "REQ-001",
                            "requirement_name": "Mobile login",
                            "description": "Users can sign in on mobile.",
                            "biz_requirement_name": "Mobile banking",
                        },
                        {
                            "requirement_id": "REQ-002",
                            "requirement_name": "Account list",
                            "description": "Users can view accounts.",
                            "biz_requirement_name": "Mobile banking",
                        },
                    ]
                },
            },
        )
    )

    assert response.success is True
    tasks = response.result["tasks"]
    project_start = date.fromisoformat("2024-01-10")
    project_end = date.fromisoformat("2024-07-10")
    for task in tasks:
        planned_start = date.fromisoformat(task["planned_start_date"])
        planned_end = date.fromisoformat(task["planned_end_date"])
        assert project_start <= planned_start <= project_end
        assert project_start <= planned_end <= project_end
        assert planned_start <= planned_end
        assert task["metadata"]["start_date"] == planned_start.strftime("%Y.%m.%d")
        assert task["metadata"]["end_date"] == planned_end.strftime("%Y.%m.%d")

    file_bytes = ArtifactExportService()._build_wbs_xlsx(response.result)
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook["WBS"]
    assert sheet.cell(row=2, column=5).value == "2024.01.10"
    assert sheet.cell(row=2, column=6).value == "2024.07.10"


def test_user_flow_output_agent_formats_generated_artifact_downloads() -> None:
    payload = ChatOutputAgent().build_display_payload(
        {
            "event": "ACTION_COMPLETED",
            "result": {
                "artifact": {
                    "artifact_id": "ART-REQ-001",
                    "artifact_type": "REQUIREMENT_SPEC",
                    "name": "요구사항명세서.xlsx",
                },
                "generated": {"artifact_type": "REQUIREMENT_SPEC"},
                "exported_file": {
                    "file_name": "요구사항명세서.xlsx",
                    "content_type": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                },
            },
        }
    )

    assert payload["state"] == "COMPLETED"
    assert payload["download_files"][0]["artifact_id"] == "ART-REQ-001"
    assert payload["download_files"][0]["file_name"].endswith(".xlsx")
    assert payload["suggested_actions"] == []
    assert_user_facing_message(payload["message"])
