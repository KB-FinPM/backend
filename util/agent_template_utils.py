# EN: Template/mapper helpers used by core artifact agents.
# KO: core artifact Agent에서 사용하는 template/mapper 로딩 유틸리티입니다.

from __future__ import annotations

import json
import os
import re
import tempfile
import unicodedata
from copy import deepcopy
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

import boto3
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl import load_workbook

from app.core.config import settings
from app.schemas.request import normalize_author_value

TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "app" / "agents" / "core_agents" / "template"
CACHE_DIR = Path(tempfile.gettempdir()) / "finpm-agent-template-cache"


def _norm(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def _decode_hash_unicode(value: str) -> str:
    return re.sub(
        r"#U([0-9A-Fa-f]{4})",
        lambda match: chr(int(match.group(1), 16)),
        value,
    )


def _relative_template_path(path_or_name: str) -> str:
    raw = str(path_or_name or "").replace("\\", "/").lstrip("/")
    if raw.startswith("template/"):
        return raw[len("template/"):]
    return raw


def _resolve_local_template_file(path_or_name: str) -> Path:
    relative = _relative_template_path(path_or_name)
    candidate = TEMPLATE_DIR / relative
    if candidate.exists():
        return candidate

    decoded_candidate = TEMPLATE_DIR / _decode_hash_unicode(relative)
    if decoded_candidate.exists():
        return decoded_candidate

    target_names = {
        _norm(Path(relative).name),
        _norm(_decode_hash_unicode(Path(relative).name)),
    }
    if TEMPLATE_DIR.exists():
        for item in TEMPLATE_DIR.rglob("*"):
            if not item.is_file():
                continue
            item_names = {
                _norm(item.name),
                _norm(_decode_hash_unicode(item.name)),
            }
            if target_names & item_names:
                return item
    raise FileNotFoundError(f"template file not found: {path_or_name}")


def _s3_client():
    kwargs: dict[str, Any] = {
        "region_name": settings.AWS_REGION,
        "verify": settings.AWS_CA_BUNDLE or settings.AWS_VERIFY_SSL,
    }
    if settings.AWS_ACCESS_KEY_ID and settings.AWS_SECRET_ACCESS_KEY:
        kwargs["aws_access_key_id"] = settings.AWS_ACCESS_KEY_ID
        kwargs["aws_secret_access_key"] = settings.AWS_SECRET_ACCESS_KEY
    return boto3.client("s3", **kwargs)


def _s3_template_key(path_or_name: str) -> str:
    relative = _relative_template_path(path_or_name)
    return f"{settings.S3_TEMPLATE_PREFIX.rstrip('/')}/{relative}".replace("//", "/")


def _download_s3_template(path_or_name: str) -> Path | None:
    if settings.S3_STORAGE_BACKEND.lower() != "s3" or not settings.S3_BUCKET_NAME:
        return None

    client = _s3_client()
    key = _s3_template_key(path_or_name)
    target = CACHE_DIR / settings.S3_BUCKET_NAME / key
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        client.download_file(settings.S3_BUCKET_NAME, key, str(target))
        return target
    except (BotoCoreError, ClientError, OSError):
        pass

    # Mac에서 업로드된 한글 파일명은 NFC/NFD가 섞일 수 있으므로,
    # prefix 목록을 조회해 정규화된 파일명이 같은 객체를 한 번 더 찾습니다.
    prefix = settings.S3_TEMPLATE_PREFIX.rstrip("/") + "/"
    target_name = _norm(Path(_relative_template_path(path_or_name)).name)
    try:
        response = client.list_objects_v2(Bucket=settings.S3_BUCKET_NAME, Prefix=prefix)
        for item in response.get("Contents", []):
            candidate_key = item.get("Key", "")
            if _norm(Path(candidate_key).name) != target_name:
                continue
            target = CACHE_DIR / settings.S3_BUCKET_NAME / candidate_key
            target.parent.mkdir(parents=True, exist_ok=True)
            client.download_file(settings.S3_BUCKET_NAME, candidate_key, str(target))
            return target
    except (BotoCoreError, ClientError, OSError):
        return None
    return None


def resolve_template_path(path_or_name: str) -> str:
    """Return a local path for a template, preferring S3 template_files.

    A mapper value like ``template/탬플릿_WBS.xlsx`` is resolved to
    ``s3://<bucket>/<S3_TEMPLATE_PREFIX>/탬플릿_WBS.xlsx`` when S3 mode is enabled.
    If download fails, the bundled ``app/agents/core_agents/template`` copy is used.
    """
    downloaded = _download_s3_template(path_or_name)
    if downloaded is not None and downloaded.exists():
        return str(downloaded)
    return str(_resolve_local_template_file(path_or_name))


def load_template_json(path_or_name: str, default: dict[str, Any] | None = None) -> dict[str, Any]:
    try:
        path = Path(resolve_template_path(path_or_name))
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return deepcopy(default or {})


def load_output_mapper() -> dict[str, Any]:
    return load_template_json("output_mapper.json", {})


def _load_local_template_json(path_or_name: str, default: dict[str, Any] | None = None) -> dict[str, Any]:
    try:
        path = _resolve_local_template_file(path_or_name)
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return deepcopy(default or {})


def load_deliverable_mapper() -> dict[str, Any]:
    mapper = load_output_mapper()
    path = get_nested(mapper, "wbs", "deliverable_mapper_path", default="deliverable_mapper.json")
    return load_template_json(path, {})


def load_deliverable_mapper_local() -> dict[str, Any]:
    mapper = _load_local_template_json("output_mapper.json", {})
    path = get_nested(mapper, "wbs", "deliverable_mapper_path", default="deliverable_mapper.json")
    return _load_local_template_json(path, {})


@lru_cache(maxsize=1)
def load_wbs_deliverable_catalog() -> list[dict[str, str]]:
    """Load the workbook-based WBS deliverable reference list."""
    path = resolve_template_path("template/산출물목록.xlsx")
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    catalog: list[dict[str, str]] = []
    current_stage = ""
    current_task = ""
    current_activity = ""

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        stage, task, activity, purpose, output = (list(row) + [None] * 5)[:5]
        if stage:
            current_stage = str(stage).strip()
        if task:
            current_task = str(task).strip()
        if activity:
            current_activity = str(activity).strip()
        deliverable = str(output or "").strip()
        if not deliverable:
            continue
        catalog.append(
            {
                "stage": current_stage,
                "task": current_task,
                "activity": current_activity,
                "purpose": str(purpose or "").strip(),
                "deliverable": deliverable,
            }
        )

    return catalog


@lru_cache(maxsize=1)
def load_wbs_deliverable_catalog_local() -> list[dict[str, str]]:
    """Load the bundled WBS deliverable reference list without S3 lookup."""
    path = TEMPLATE_DIR / "산출물목록.xlsx"
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    catalog: list[dict[str, str]] = []
    current_stage = ""
    current_task = ""
    current_activity = ""

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        stage, task, activity, purpose, output = (list(row) + [None] * 5)[:5]
        if stage:
            current_stage = str(stage).strip()
        if task:
            current_task = str(task).strip()
        if activity:
            current_activity = str(activity).strip()
        deliverable = str(output or "").strip()
        if not deliverable:
            continue
        catalog.append(
            {
                "stage": current_stage,
                "task": current_task,
                "activity": current_activity,
                "purpose": str(purpose or "").strip(),
                "deliverable": deliverable,
            }
        )

    return catalog


def load_wbs_template() -> dict[str, Any]:
    return _load_local_template_json("wbs_template.json", {"common_items": []})


def load_wbs_common_rows(
    *,
    template_path: str | None = None,
    sheet_name: str = "WBS",
    start_row: int = 2,
    end_row: int | None = 86,
) -> list[dict[str, Any]]:
    """Load the static WBS prefix rows from the bundled JSON template."""
    template = load_wbs_template()
    common_items = template.get("common_items", [])
    rows: list[dict[str, Any]] = []
    max_count = max((end_row or start_row) - start_row + 1, 0) if end_row is not None else len(common_items)
    for item in common_items[:max_count]:
        if not isinstance(item, dict):
            continue
        wbs_name = str(item.get("wbs_name") or "").strip()
        if not wbs_name:
            continue
        rows.append(
            {
                "level": "" if item.get("level") is None else str(item.get("level")),
                "wbs_id": "" if item.get("wbs_id") is None else str(item.get("wbs_id")),
                "wbs_name": wbs_name,
                "deliverable": "" if item.get("deliverable") is None else str(item.get("deliverable")),
            }
        )
    return rows


def get_nested(config: dict[str, Any], *keys: str, default: Any = None) -> Any:
    current: Any = config
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return default
        current = current[key]
    return current


def build_template_context(project_id: str = "", context: dict[str, Any] | None = None) -> dict[str, str]:
    context = context or {}
    author = normalize_author_value(context.get("author")) or normalize_author_value(
        context.get("writer")
    )
    return {
        "project_id": str(project_id or ""),
        "project_name": str(context.get("project_name") or context.get("project_nm") or "프로젝트명"),
        "author": author,
        "today": datetime.today().strftime("%Y-%m-%d"),
    }


def get_value(source: Any, field_expr: Any, *, context: dict[str, str] | None = None, row_number: int | None = None) -> Any:
    if field_expr is None:
        return ""
    if not isinstance(field_expr, str):
        return field_expr
    if field_expr == "":
        return ""
    if field_expr == "row_number":
        return row_number or ""
    if context and field_expr in context:
        return context.get(field_expr, "")

    if isinstance(source, dict):
        metadata = source.get("metadata") or {}
    else:
        metadata = getattr(source, "metadata", {}) or {}

    for raw_field in field_expr.split("|"):
        field = raw_field.strip()
        if not field:
            continue
        if context and field in context:
            value = context.get(field, "")
        elif isinstance(source, dict):
            value = source.get(field, "")
            if value in (None, ""):
                value = metadata.get(field, "") if isinstance(metadata, dict) else ""
        else:
            value = getattr(source, field, "")
            if value in (None, "") and isinstance(metadata, dict):
                value = metadata.get(field, "")
        if value not in (None, ""):
            return value
    return ""


def build_placeholder_values(placeholders: dict[str, str], source: Any = None, context: dict[str, str] | None = None) -> dict[str, str]:
    context = context or {}
    return {str(k): str(get_value(source or {}, v, context=context) or "") for k, v in (placeholders or {}).items()}


def mapper_summary_for_prompt() -> str:
    mapper = load_output_mapper()
    req_columns = get_nested(mapper, "requirement_spec", "data_sheet", "columns", default=[])
    wbs_mapper = get_nested(mapper, "wbs", "data_sheet", "columns", default=[])
    return json.dumps(
        {
            "requirement_spec_columns": req_columns,
            "wbs_columns": wbs_mapper,
            "screen_plan": get_nested(mapper, "screen_plan", default={}),
        },
        ensure_ascii=False,
        indent=2,
    )


def output_file_name(document_key: str, fallback: str) -> str:
    mapper = load_output_mapper()
    doc = get_nested(mapper, "output_files", "documents", document_key, default={})
    name = doc.get("document_name") if isinstance(doc, dict) else None
    ext = doc.get("extension") if isinstance(doc, dict) else None
    if name and ext:
        return f"{name}{ext}"
    return fallback


def find_deliverable(name: str, phase: str, deliverable_mapper: dict[str, Any] | None = None) -> str:
    text = f"{name} {phase}"
    catalog = load_wbs_deliverable_catalog()
    matches: list[str] = []
    for item in catalog:
        haystack = " ".join(
            value
            for value in (
                item.get("stage", ""),
                item.get("task", ""),
                item.get("activity", ""),
                item.get("purpose", ""),
                item.get("deliverable", ""),
            )
            if value
        )
        if not haystack:
            continue
        if str(phase or "").strip() and str(phase).strip() not in haystack:
            continue
        if not any(str(token).lower() in text.lower() for token in (item.get("task", ""), item.get("activity", ""), item.get("deliverable", "")) if token):
            continue
        deliverable = item.get("deliverable", "")
        if deliverable and deliverable not in matches:
            matches.append(deliverable)
    if matches:
        return ", ".join(matches[:3])

    mapper = deliverable_mapper or load_deliverable_mapper()
    for rule in mapper.get("keyword_rules", []):
        keywords = rule.get("keywords") or []
        if any(str(keyword).lower() in text.lower() for keyword in keywords):
            deliverables = rule.get("deliverables") or []
            return ", ".join(str(item) for item in deliverables[:2] if item)
    for key, value in (mapper.get("default_by_phase") or {}).items():
        if str(key) in str(phase):
            return str(value)
    return ""
